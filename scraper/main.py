import openpyxl
import requests
from bs4 import BeautifulSoup
import json
import os
import camelot


def get_column_with_name(col_name):
    """
    XLSX from https://www.ema.europa.eu/en/medicines/download-medicine-data

    :param col_name: Name of column to retrieve from .xlsx file
    :return: column values
    """

    # Open the Workbook
    workbook = openpyxl.load_workbook("data/Medicines_output_herbal_medicines.xlsx")

    # Define variable to read the active sheet:
    worksheet = workbook.active

    url_col = None
    # Iterate over the column names (8th row) to find the column with URLs
    for col in worksheet.iter_cols(1, worksheet.max_column):
        if col[8].value == col_name:
            url_col = col
            break

    if not url_col:
        raise ValueError(f"No column with name {col_name}!")

    return url_col


def get_the_assessment_pdfs():
    """
    Get from the webpages from XLSX file. Get each of webpages HTML, and filter it
    to get the links to pdf files. Save list of links to file.
    """
    pdfs = {}
    herb_urls = []

    # Get page url
    url_col = get_column_with_name("URL")
    for url_row in url_col:
        if str(url_row.value).startswith("http"):
            herb_urls.append(url_row.value)

    for url in herb_urls:

        # Get page
        page = requests.get(url)

        # Preapre soup (mmm..)
        data = page.text
        soup = BeautifulSoup(data, features="html.parser")

        # Find the page title and remove suffix
        title = soup.find("title").string.rstrip(" | European Medicines Agency")

        pdfs[title] = []

        # Find the URLs to pdfs with assessment-reports
        for li in soup.find_all(class_="ecl-list-item ema-list-item ema-list-item--file"):
            pdf_url = li.find('a')['href']
            if "assessment-report" in pdf_url:
                print(title + " " + pdf_url)
                pdfs[title].append(pdf_url)

    # backup, bo nie wiem, zaraz mnie zbanuja za crawlowanie
    with open('assesment_pdfs.json', 'w+') as file:
        json.dump(pdfs, file)

    print("ok")


def find_best_assessment(docs):
    """
    :param docs: List of urls to pdf files
    :return: Link to url that I like the most
    """
    # Clean up reference documents
    docs = [doc for doc in docs if "herbal-references" not in doc]

    # Clean up superseded documents
    docs = [doc for doc in docs if "superseded" not in doc]

    # Clean up addendums- even though, maybe we shouldn't
    # TODO
    docs = [doc for doc in docs if "addendum" not in doc]

    # remove drafts if not all of them are drafts
    if len(docs) > 1:
        draft_docs = [doc for doc in docs if "draft" in doc]
        if 0 < len(draft_docs) < len(docs):
            docs = [doc for doc in docs if "draft" not in doc]

    # wops, something still was not deleted
    if len(docs) > 1:
        print(docs)

    if len(docs) == 0:
        return None

    return docs[0]


def download_assessment_pdfs():
    print("load")
    with open("assesment_pdfs.json", "r") as file:
        assessments = json.load(file)

    for herb, docs in assessments.items():
        # get one assesment out of list of it
        doc = find_best_assessment(docs)
        if not doc:
            print(f"{herb} does not have assessment")
            continue
        # name of herb as name of directory- no spaces
        herb = herb.replace(" ", "_")

        # create dir if not exists
        try:
            os.mkdir(f"data/{herb}")
        except Exception:
            continue

        # download herb docs
        print(f"Downloading the doc for herb {herb}")
        pdf_resp = requests.get(doc)
        with open(f"data/{herb}/doc.pdf", "wb") as file:
            file.write(pdf_resp.content)


if __name__ == "__main__":
    # Executed that once and did backup, in order to not fucking crawl through everything on each run
    #get_the_assessment_pdfs()

    # Downloading assesment_pdfs.json
    #assesment_pdfs = download_assessment_pdfs()

    # extract all the tables in the PDF file
    abc = camelot.read_pdf("data/Absinthii_herb/doc.pdf")  # address of file location

    # print the first table as Pandas DataFrame
    print(abc[0].df)

    print("ok")
